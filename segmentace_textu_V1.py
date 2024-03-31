import cv2
import numpy as np
from pdf2image import convert_from_path
from skimage.filters import threshold_sauvola
from openpyxl import Workbook
import os
import capture_tables as f
from sklearn.cluster import MeanShift
import pandas as pd
import pytesseract

os.system('rm -rf excel*')

def segment_text(images, sauvola_window_size, bandwidths, area, bandwidth):
    try:
        for page_num, image in enumerate(images):
            img = np.array(image)
            gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
            gray = cv2.GaussianBlur(gray, (5, 5), 0)

            thresh_sauvola = threshold_sauvola(gray, window_size=sauvola_window_size)
            binary_sauvola = gray > thresh_sauvola
            binary_sauvola = (binary_sauvola * 255).astype(np.uint8)
            binary_sauvola = cv2.bitwise_not(binary_sauvola)

            horizontal_lines, vertical_lines = f.find_table_structure(binary_sauvola)
            table_structure = cv2.bitwise_or(horizontal_lines, vertical_lines)

            contours, hierarchy = cv2.findContours(table_structure, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

            contours = sorted(contours, key=lambda x: x[0][0][1])

            # Assuming the largest contour is the table we want to align
            index = 0
            for contour in contours:
                # Find the bounding rectangle and crop the region from table_structure
                if cv2.contourArea(contour) > area:
                    index += 1
                    wb = Workbook()
                    ws = wb.active
                    x, y, w, h = cv2.boundingRect(contour)
                    cropped_table = table_structure[y:y + h, x:x + w]
                    cropped_img = binary_sauvola[y:y + h, x:x + w]

                    # Assuming the contour can be approximated to a rectangle for perspective transform
                    rect = cv2.minAreaRect(contour)
                    box = cv2.boxPoints(rect)
                    box = np.array(box, dtype='float32')

                    # Translate box points based on crop
                    box[:, 0] -= x
                    box[:, 1] -= y

                    # Apply perspective transform
                    corrected_table, corrected_img = f.apply_perspective_transform(cropped_table, cropped_img, box)

                    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))

                    # Perform dilation
                    corrected_table = cv2.dilate(corrected_table, kernel)


                    corrected_img_rgb = cv2.cvtColor(corrected_img, cv2.COLOR_GRAY2BGR)

                    cv2.imwrite(f"output_masks/page_{page_num + 1}_contour_{index}.png", corrected_table)
                    cv2.imwrite(f"output/page_{page_num + 1}_contour_{index}.png", corrected_img)

                    contours, hierarchy = cv2.findContours(corrected_table, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)


                    height, width = corrected_table.shape[:2]

                    extracted_texts = []
                    corners_xs = []
                    corners_ys = []
                    data = []

                    # Assuming 'contours' and 'hierarchy' are defined
                    for i, contour in enumerate(contours):
                        if hierarchy[0][i][3] != -1:
                            x, y, w, h = cv2.boundingRect(contour)
                            if x > 0 and y > 0 and (x + w) < width and (y + h) < height:
                                cv2.rectangle(corrected_img_rgb, (x, y), (x + w, y + h), (0, 0, 255), 2)
                                cropped_img = cv2.bitwise_not(corrected_img[y:y + h, x:x + w])

                                extracted_text = pytesseract.image_to_string(cropped_img, config='--psm 6', lang='eng+ara')
                                topLeftCorner = [x, y]  # Using center point for clustering
                                if extracted_text:  # Only add if text exists
                                    data.append((topLeftCorner, extracted_text))
                                else:
                                    data.append((topLeftCorner, ""))


                    # Separate coordinates and texts for clustering
                    for item in data:
                        corners_xs.append(item[0][0])
                        corners_ys.append(item[0][1])
                        extracted_texts.append(item[1])

                    # Apply Mean Shift for row and column clustering
                    ms_x = MeanShift(bandwidth=bandwidth, bin_seeding=False)
                    ms_y = MeanShift(bandwidth=bandwidth, bin_seeding=False)

                    ms_x.fit(np.array(corners_xs).reshape(-1, 1))
                    ms_y.fit(np.array(corners_ys).reshape(-1, 1))

                    # Obtain and sort the cluster centers
                    sorted_row_indices = np.argsort(ms_y.cluster_centers_.ravel())
                    sorted_col_indices = np.argsort(ms_x.cluster_centers_.ravel())

                    # Initialize Excel workbook and sheet
                    wb = Workbook()
                    ws = wb.active

                    # Create a matrix for texts with dimensions based on the sorted clusters
                    text_matrix = [["" for _ in range(len(sorted_col_indices))] for _ in range(len(sorted_row_indices))]

                    # Map each text to its cell in the matrix based on the nearest cluster center
                    for (coords, text) in data:
                        x, y = coords
                        col_cluster = np.argmin(np.abs(ms_x.cluster_centers_ - x))
                        row_cluster = np.argmin(np.abs(ms_y.cluster_centers_ - y))
                        col = np.where(sorted_col_indices == col_cluster)[0][0]
                        row = np.where(sorted_row_indices == row_cluster)[0][0]
                        text_matrix[row][col] = text

                    # Fill the Excel sheet with the text matrix
                    for i, row_content in enumerate(text_matrix):
                        for j, cell_text in enumerate(row_content):
                            ws.cell(row=i + 1, column=j + 1, value=str(cell_text))
                            
                    if not os.path.exists('excel'):
                        os.makedirs('excel')
                    excel_path = f"excel/page_{page_num + 1}_contour_{index}.xlsx"
                    wb.save(excel_path)
                    print(f"{excel_path} saved")
        return True
    except Exception as e:
        print(e)
        return False

def main(pdf_path):

    sauvola_window_size = 1101
    bandwidths = [12]
    area = 15000
    bandwidth = 12
    images = convert_from_path(pdf_path, dpi=300)
    return segment_text(images, sauvola_window_size, bandwidths, area, bandwidth)

